import numpy as np
import matplotlib.pyplot as plt
import openpyxl

# Computing a meshgrid
dt = 0.1
x = np.arange(-2, 2, dt)
y = np.arange(-2, 2, dt)
xx, yy = np.meshgrid(x, y)

m, n = xx.shape

states = np.zeros([m*n, 2])

for i in range(0, m):
    for j in range(0, n):
        counter = i*m + j
        states[counter, :] = np.array([xx[i, j], yy[i, j]])
        

# Working with excel files
wb = openpyxl.load_workbook('example.xlsx')
if 'States' in wb.sheetnames:
    del wb['States']

wb.create_sheet(index = 0, title = 'States')

if 'Time' in wb.sheetnames:
    del wb['Time']

wb.create_sheet(index = 1, title = 'Time')


if 'Data' in wb.sheetnames:
    del wb['Data']

wb.create_sheet(index = 2, title = 'Data')

state_sheet = wb['States']
time_sheet = wb['Time']
data_sheet = wb['Data']
counter = 0
for i in range(0, m):
    for j in range(0, n):
        state_sheet['A'+str(counter+1)] = states[counter, 0]
        state_sheet['B'+str(counter+1)] = states[counter, 1]
        time_sheet['A'+str(counter+1)] = 'None'
        counter += 1

data_sheet['A1'] = m
data_sheet['B1'] = n

wb.save('./example.xlsx')

